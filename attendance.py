import cv2
import numpy as np
import face_recognition
import os
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import simpledialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ================= CONFIG =================
IMAGE_PATH = "images"
EXCEL_FILE = "Attendance.xlsx"
TEACHER_PASSWORD = "admin123"

PRESENT_HOUR = 9
LATE_LIMIT_MINUTE = 15

# ================= LOAD STUDENT IMAGES =================
images = []
student_names = []

for file in os.listdir(IMAGE_PATH):
    img = cv2.imread(os.path.join(IMAGE_PATH, file))
    images.append(img)
    student_names.append(os.path.splitext(file)[0])

# ================= ENCODE FACES =================
def find_encodings(images):
    encodings = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        enc = face_recognition.face_encodings(img)
        if enc:
            encodings.append(enc[0])
    return encodings

encode_list_known = find_encodings(images)

# ================= TKINTER ROOT (CENTERED) =================
root = tk.Tk()
root.title("Smart Attendance Dashboard")

W, H = 420, 500
sw = root.winfo_screenwidth()
sh = root.winfo_screenheight()
x = (sw // 2) - (W // 2)
y = (sh // 2) - (H // 2)

root.geometry(f"{W}x{H}+{x}+{y}")
root.resizable(False, False)

# ================= TEACHER PASSWORD =================
def ask_teacher_permission(name):
    password = simpledialog.askstring(
        "Teacher Approval",
        f"Enter password for {name}:",
        show="*"
    )
    return password == TEACHER_PASSWORD

# ================= CENTER OPENCV WINDOW =================
def center_opencv_window(name, width=800, height=600):
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x = (sw // 2) - (width // 2)
    y = (sh // 2) - (height // 2)

    cv2.namedWindow(name, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(name, width, height)
    cv2.moveWindow(name, x, y)

# ================= EXCEL FORMATTING (WITH SUNDAY) =================
def format_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    year = datetime.now().year

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 12

        if col != 1:
            try:
                date_obj = datetime.strptime(f"{cell.value}-{year}", "%d-%b-%Y")
                if date_obj.weekday() == 6:  # Sunday
                    for r in range(1, max_row + 1):
                        ws.cell(row=r, column=col).fill = PatternFill(
                            "solid", fgColor="FF9999"
                        )
            except:
                pass

    for r in range(2, max_row + 1):
        for c in range(2, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="center")

            if cell.value == "P":
                cell.fill = PatternFill("solid", fgColor="90EE90")
            elif cell.value == "Late":
                cell.fill = PatternFill("solid", fgColor="FFD966")
            elif cell.value == "A":
                cell.fill = PatternFill("solid", fgColor="F4CCCC")

    wb.save(EXCEL_FILE)

# ================= MARK ATTENDANCE =================
def mark_attendance(name):
    now = datetime.now()
    today = now.strftime('%d-%b')
    current_time = now.time()

    df = pd.read_excel(EXCEL_FILE)
    df.columns = [col.strftime('%d-%b') if isinstance(col, datetime) else col for col in df.columns]

    if today not in df.columns or name not in df['Name/Date'].values:
        return

    row = df[df['Name/Date'] == name].index[0]
    df[today] = df[today].astype(str).replace("nan", "")

    present_time = now.replace(hour=PRESENT_HOUR, minute=0, second=0).time()
    late_limit = now.replace(hour=PRESENT_HOUR, minute=LATE_LIMIT_MINUTE, second=0).time()

    if df.at[row, today] in ["P", "Late"]:
        return

    if current_time < present_time:
        df.at[row, today] = "P"
    elif present_time <= current_time <= late_limit:
        df.at[row, today] = "Late"
    else:
        if ask_teacher_permission(name):
            df.at[row, today] = "Late"
        else:
            df.at[row, today] = "A"

    df.to_excel(EXCEL_FILE, index=False)
    format_excel()

# ================= AUTO ABSENT =================
def mark_absent_students():
    today = datetime.now().strftime('%d-%b')
    df = pd.read_excel(EXCEL_FILE)
    df.columns = [col.strftime('%d-%b') if isinstance(col, datetime) else col for col in df.columns]

    if today not in df.columns:
        return

    df[today] = df[today].astype(str).replace("nan", "")

    for i in range(len(df)):
        if df.at[i, today] == "":
            df.at[i, today] = "A"

    df.to_excel(EXCEL_FILE, index=False)
    format_excel()

# ================= MONTHLY PERCENTAGE =================
def show_monthly_percentage():
    df = pd.read_excel(EXCEL_FILE)
    df.columns = [col.strftime('%d-%b') if isinstance(col, datetime) else col for col in df.columns]

    month = datetime.now().strftime('%b')
    year = datetime.now().year

    date_cols = []
    for col in df.columns[1:]:
        try:
            d = datetime.strptime(f"{col}-{year}", "%d-%b-%Y")
            if d.strftime('%b') == month and d.weekday() != 6:
                date_cols.append(col)
        except:
            pass

    if not date_cols:
        messagebox.showinfo("Monthly %", "No data for this month")
        return

    result = ""
    for i in range(len(df)):
        attended = sum(df.loc[i, date_cols].isin(["P", "Late"]))
        percentage = (attended / len(date_cols)) * 100
        result += f"{df.at[i,'Name/Date']}: {percentage:.2f}%\n"

    messagebox.showinfo("Monthly Attendance %", result)

# ================= SUMMARY =================
def show_summary():
    today = datetime.now().strftime('%d-%b')
    df = pd.read_excel(EXCEL_FILE)
    df.columns = [col.strftime('%d-%b') if isinstance(col, datetime) else col for col in df.columns]

    if today not in df.columns:
        return

    messagebox.showinfo(
        "Summary",
        f"Present: {(df[today]=='P').sum()}\n"
        f"Late: {(df[today]=='Late').sum()}\n"
        f"Absent: {(df[today]=='A').sum()}"
    )

# ================= CAMERA =================
def start_camera():
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    center_opencv_window("Face Recognition")

    while True:
        success, img = cap.read()
        if not success:
            break

        img_small = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        img_small = cv2.cvtColor(img_small, cv2.COLOR_BGR2RGB)

        faces = face_recognition.face_locations(img_small)
        encodes = face_recognition.face_encodings(img_small, faces)

        for encode_face in encodes:
            distances = face_recognition.face_distance(encode_list_known, encode_face)
            matches = face_recognition.compare_faces(encode_list_known, encode_face)

            if len(distances) == 0:
                continue

            idx = np.argmin(distances)
            if matches[idx]:
                mark_attendance(student_names[idx])

        cv2.imshow("Face Recognition", img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    mark_absent_students()

# ================= DASHBOARD =================
tk.Label(root, text="Smart Attendance System",
         font=("Arial", 16, "bold")).pack(pady=15)

tk.Button(root, text="📷 Start Attendance",
          width=30, command=start_camera).pack(pady=8)

tk.Button(root, text="📊 Attendance Summary",
          width=30, command=show_summary).pack(pady=8)

tk.Button(root, text="📈 Monthly Attendance %",
          width=30, command=show_monthly_percentage).pack(pady=8)

tk.Button(root, text="❌ Exit",
          width=30, command=root.destroy).pack(pady=15)

root.mainloop()
